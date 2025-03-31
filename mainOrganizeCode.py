# Annie Jaynes, Blake Pead, Joaquin Elizalde, Michael Jones, Becca Braatz
# This program cleans up some data from an Excel workbook by reformatting and summarizing the data.

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

# Loads unorganized data
myWorkbook = openpyxl.load_workbook('Poorly_Organized_Data_1.xlsx')
newWorkbook = Workbook()

# Activates the unorganized data worksheet to start grabbing data
currSheet = myWorkbook.active

# Creates 'current class' variable
current_class = 'None'

# Loop through each student
for row in currSheet.iter_rows(min_row=2,values_only=True):

    # Check to see if we are working with the right class for the student
    if row[0] != current_class:
        current_class = row[0]

        # Create new worksheets for each class
        newWorkbook.create_sheet(current_class)
        # In each sheet, create columns for last name, first name, student ID, and grade with the student data for that class placed there
        newWorkbook[current_class]['A1'] = 'Last Name'
        newWorkbook[current_class]['B1'] = 'First Name'
        newWorkbook[current_class]['C1'] = 'Student ID'
        newWorkbook[current_class]['D1'] = 'Grade'

    # For each student in our original sheet, split the data into a list and add it into the proper row of our new data
    stud_data = (row[1].split('_'))
    stud_data.append(row[2])
    newWorkbook[current_class].append(stud_data)

# Creates a bold font
f1 = Font(bold=True)

# Removes the original 'Sheet' sheet
newWorkbook.remove(newWorkbook['Sheet'])

# Loops through each of our sheets that we have created
for sheet in newWorkbook:
    # Creates each of our summary titles
    sheet['F1'] = 'Summary Statistics'
    sheet['F2'] = 'Highest Grade'
    sheet['F3'] = 'Lowest Grade'
    sheet['F4'] = 'Mean Grade'
    sheet['F5'] = 'Median Grade'
    sheet['F6'] = 'Students In Class'

    # Gets the range of all of the data that we have created
    lastRow = sheet.max_row

    # Functions to get the values of our summary
    sheet['G1'] = 'Value'
    sheet['G2'] = f'=MAX(D2:D{lastRow})'
    sheet['G3'] = f'=MIN(D2:D{lastRow})'
    sheet['G4'] = f'=AVERAGE(D2:D{lastRow})'
    sheet['G5'] = f"=MEDIAN(D2:D{lastRow})"   
    sheet['G6'] = f'=COUNT(D2:D{lastRow})'

    # Applies bold font created earlier
    for cell in sheet['A1:G1'][0]:
        cell.font = f1

    # Add filter to each of the first four columns in each sheet
    sheet.auto_filter.ref = f"A1:D{lastRow}"

    # Adjust width of columns
    sheet.column_dimensions["A"].width = len(sheet["A1"]) + 5
    sheet.column_dimensions["B"].width = len(sheet["B1"]) + 5
    sheet.column_dimensions["C"].width = len(sheet["C1"]) + 5
    sheet.column_dimensions["D"].width = len(sheet["D1"]) + 5
    sheet.column_dimensions["E"].width = len(sheet["E1"]) + 5
    sheet.column_dimensions["F"].width = len(sheet["F1"]) + 5
    sheet.column_dimensions["G"].width = len(sheet["G1"]) + 5

# Removes the original 'Sheet' sheet
newWorkbook.remove(newWorkbook['Sheet'])

# Save our organized data
newWorkbook.save(filename='filename.xlsx')

# Closes both workbooks
myWorkbook.close()
newWorkbook.close()